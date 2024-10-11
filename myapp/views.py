from .forms import TextInputForm
from django.shortcuts import render, redirect
from django.contrib import messages
from collections import defaultdict
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Font, Alignment, Border, Side
import os, re, io, docx, PyPDF2, pdfplumber
from django.core.files.uploadedfile import UploadedFile
import pandas as pd
import asyncio

def export_to_excel(request):

    word_count = request.session.get('word_count', None)

    if not word_count:
        return HttpResponse("No data to export", status=400)
    
    data = [(index + 1, word, count) for index, (word, count) in enumerate(word_count.items())]
    
    df = pd.DataFrame(data, columns=['№', 'Термин', 'Жиілігі'])
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Терминдер', startrow=2, index=False)

        worksheet = writer.sheets['Терминдер']

        format_excel_file(worksheet, word_count)
    
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export-termins.xlsx'
    
    return response

def format_excel_file(worksheet, word_count):
    worksheet.merge_cells('A1:C1')
    worksheet.merge_cells('A2:C2')

    worksheet['A1'] = 'Термин сөздердің жиілігі'
    worksheet['A2'] = 'Барлық термин сөздердің саны: ' + str(len(word_count))

    left_alignment = Alignment(horizontal='left', vertical='center')
    indent_alignment = Alignment(horizontal='left', indent=2)
    indent_right_alignment = Alignment(horizontal='right', indent=2)
    center_alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A1'].alignment = left_alignment
    worksheet['A2'].alignment = left_alignment

    fontBold = Font(name='Open Sans', size=14, bold=True)
    fontLight = Font(name='Open Sans', size=14, bold=False)
    worksheet['A1'].font = fontBold
    worksheet['A2'].font = fontLight

    worksheet.column_dimensions['A'].width = 7 
    worksheet.column_dimensions['B'].width = 25 
    worksheet.column_dimensions['C'].width = 12 

    worksheet.row_dimensions[1].height = 25 
    worksheet.row_dimensions[2].height = 25 
    worksheet.row_dimensions[3].height = 20

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = indent_alignment 

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = indent_right_alignment 

    general_font = Font(name='Open Sans', size=13)
    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.font = general_font
            cell.border = thin_border

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = center_alignment

    header_font = Font(name='Open Sans', size=13, bold=True)
    for cell in worksheet[3]:
        cell.font = header_font 

def extract_text_from_docx(file_path):
    doc = docx.Document(file_path)
    full_text = []
    for paragraph in doc.paragraphs:
        full_text.append(paragraph.text)
    return '\n'.join(full_text)

# def extract_text_from_pdf(file_path):
#     text = ""
#     pdf_reader = PyPDF2.PdfFileReader(file_path)
#     full_text = []
#     for page_num in range(pdf_reader.numPages):
#         page = pdf_reader.getPage(page_num)
#         full_text.append(page.extract_text())
#     return '\n'.join(full_text)

# def extract_text_from_pdf(file):
#     with pdfplumber.open(file) as pdf:
#         full_text = []
#         for page in pdf.pages:
#             full_text.append(page.extract_text())
#         return '\n'.join(full_text)

def extract_terms_with_patterns_from_file(text, start_patterns=None, middle_patterns=None, end_patterns=None):

    words = re.findall(r'\b\w+\b', text)

    word_count = defaultdict(int)
    
    matching_words = []
    
    for word in words:

        if start_patterns:
            starts_with = any(word.startswith(pattern) for pattern in start_patterns)
        else:
            starts_with = True
        
        if middle_patterns:
            contains_middle = any(pattern in word for pattern in middle_patterns)
        else:
            contains_middle = True
        
        if end_patterns:
            ends_with = any(word.endswith(pattern) for pattern in end_patterns)
        else:
            ends_with = True 
        
        if starts_with or contains_middle or ends_with:
            # matching_words.append(word)
            word_count[word] += 1

    return word_count

def read_pdf_in_chunks(uploaded_file: UploadedFile):
    file_data = b''  # Collect the raw binary data
    for chunk in uploaded_file.chunks():
        file_data += chunk  # Append each chunk to the file_data variable
    return file_data

def extract_text_from_pdf(uploaded_file):
    # Step 1: Read the PDF file in chunks
    pdf_data = read_pdf_in_chunks(uploaded_file)

    # Step 2: Process the PDF using pdfplumber
    extracted_text = []
    
    with pdfplumber.open(io.BytesIO(pdf_data)) as pdf:
        total_pages = len(pdf.pages)  # Get the total number of pages
        print(f"Total pages: {total_pages}")
        
        # Process each page one by one
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if text:
                extracted_text.append(text)

            # Print progress for debugging large PDFs
            print(f"Processed page {page_number} of {total_pages}")
    
    return '\n'.join(extracted_text)  # Join all extracted text into a single string

def extract_text_from_large_pdf(uploaded_file, batch_size=100):
    pdf_data = read_pdf_in_chunks(uploaded_file)
    extracted_text = []

    with pdfplumber.open(io.BytesIO(pdf_data)) as pdf:
        total_pages = len(pdf.pages)
        print(f"Total pages: {total_pages}")

        # Batch processing
        for i in range(0, total_pages, batch_size):
            batch_pages = pdf.pages[i:i + batch_size]  # Get the next batch of pages
            print(f"Processing batch {i + 1} to {min(i + batch_size, total_pages)}")

            for page in batch_pages:
                text = page.extract_text()
                if text:
                    extracted_text.append(text)

            # After processing the batch, you can free up memory (if needed)
            del batch_pages

    return '\n'.join(extracted_text)
        
def text_input_view(request):
    start_patterns = ['анти', 'ам', 'тер', 'алг', 'дис', 'дез', 'контр', 'транс', 'супер', 'пан']
    middle_patterns = ['тр', 'гр', 'тика', 'тив', 'тр', 'зит']
    end_patterns = ['ка', 'ика', 'тандыру', 'ландыру', 'дандыру', 'ль', 'из', 'нт', 'гат', 'ив', 'азм', 'сив', 'ив', 'бент', 'ия', 'ция', 'лятор', 'ин', 'аль', 'иль', 'итм', 'бра', 'оль', 'иф', 'оф', 'ид', 'тив', 'виз', 'фа', 'ний', 'зур', 'ика', 'ик', 'ид', 'ин', 'икс', 'ция', 'иту', 'рм', 'из', 'изм', 'ид', 'цев', 'аль', 'иор', 'ент', 'ия', 'фер', 'ом', 'рил', 'игн', 'оф', 'опф', 'инг', 'ип', 'ейн', 'етр', 'аф', 'иг', 'ель', 'ифм', 'метр', 'пнев', 'вив', 'иля', 'ит']

    if request.method == 'POST':
        form = TextInputForm(request.POST, request.FILES)
        if form.is_valid():
            manual_text = form.cleaned_data.get('text_area')

            uploaded_file = form.cleaned_data.get('text_file')
            
            file_text = ""
            if uploaded_file:
                file_name, file_extension = os.path.splitext(uploaded_file.name)

                if file_extension == '.txt':
                    file_text = uploaded_file.read().decode('utf-8')
                elif file_extension == '.docx':
                    file_text = extract_text_from_docx(uploaded_file)
                elif file_extension == '.pdf':
                    file_text = extract_text_from_large_pdf(uploaded_file, batch_size=100)
                else:
                    raise ValueError(f"Unsupported file type: {file_extension}")
                
            combined_text = manual_text + file_text

            word_count = extract_terms_with_patterns_from_file(combined_text, start_patterns, middle_patterns, end_patterns)

            request.session['word_count'] = word_count

            # return JsonResponse({'message': 'File uploaded successfully'})
            
            word_count_list = sorted(word_count.items(), key=lambda item: item[1], reverse=True)  # Sort by frequency descending

            return render(request, 'result.html', {'word_count_list': word_count_list})

    else:
        form = TextInputForm()

    return render(request, 'form.html', {'form': form})

def result_view(request):
    
    print("Session data: ", request.session.get('word_count'))
    
    word_count = request.session.get('word_count', None)

    if not word_count:
        # If no word count data is available, show an error or redirect
        return render(request, 'error.html', {'error': 'No data found. Please upload a file and process it first.'})

    # Convert word_count dictionary to a list of tuples [(word, count), ...]
    word_count_list = sorted(word_count.items(), key=lambda item: item[1], reverse=True)  # Sort by frequency descending

    # Pass the word_count list to the result template
    return render(request, 'result.html', {'word_count_list': word_count_list})

def home(request):
    return redirect('text_input')